from django.shortcuts import render
from rest_framework.decorators import api_view
from django.http import HttpRequest
from django.http import JsonResponse
import socket
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from User.models import User
from .models import FichePresnce
from .Serializers import FichePresnceSerializer
import ipaddress
from django.core.mail import send_mail
import socket
import requests
from django.http import JsonResponse

@api_view(['GET'])
def my_view(request):
    # Récupérer l'adresse IP locale (privée)
    hostname = socket.gethostname()
    local_ip_address = socket.gethostbyname(hostname)

    # Récupérer l'adresse IP publique
    public_ip_address = requests.get('https://api.ipify.org').text

    # Assurez-vous que vous avez récupéré les deux adresses IP
    if local_ip_address and public_ip_address:
        # Construisez une réponse JSON avec les adresses IP
        data = {
            'private_ip_address': local_ip_address,
            'ip_address': public_ip_address
        }
        return JsonResponse(data)
    else:
        # Si l'une des adresses IP n'est pas disponible, retournez une erreur
        return JsonResponse({'error': 'Adresse IP non disponible'}, status=400)
  
# Fonction pour vérifier l'adresse IP


import ipaddress

def is_ip_allowed(ip_address):
    try:
        # Plages d'adresses IP autorisées
        start_ip_1 = ipaddress.ip_address('41.229.132.70')
        end_ip_1 = ipaddress.ip_address('41.229.132.125')
        specific_ip_2 = ipaddress.ip_address('41.229.132.66')
        
        ip = ipaddress.ip_address(ip_address)
        
        return (start_ip_1 <= ip <= end_ip_1) or (ip == specific_ip_2)
    except ValueError:
        return False  
    

@api_view(['GET'])
def check_ip_allowed(request, ip_address):
    is_allowed = is_ip_allowed(ip_address)
    return JsonResponse({'allowed': is_allowed})

@api_view(['POST'])
def create_groupe(request):
    if request.method == 'POST':
        groupe_serializer = FichePresnceSerializer(data=request.data)
        if groupe_serializer.is_valid():
            groupe = groupe_serializer.save()

            # Envoyer un e-mail à chaque étudiant dans le groupe
            for etudiant in groupe.Etudiants:
                user_cin = etudiant.Cin
                try:
                    user = User.objects.get(cin=user_cin)
                    user_email = user.email
                except User.DoesNotExist:
                    return Response({"message": "User not found"}, status=status.HTTP_404_NOT_FOUND)

                subject = ' Etat d\'abscence'
                message = f'Bonjour {etudiant.NomEtudiant},\n\n Vous avez été noté {etudiant.etatAbscence}  lors de la dernière séance de cours..\n\n Détails du Séance :\n Nom du groupe : {groupe.nomGroupe}\nSeance : {groupe.Seance} \n Date de la séance : {groupe.DateSeance} \n Sale de la séance : {groupe.Sale} \n Matière : {groupe.Matiere}'
                send_mail(
                    subject,
                    message,
                    'adminmira@coursenligne.info',  # Remplacez par votre adresse e-mail
                    [user_email],  # Assurez-vous que votre modèle Etudiant a un champ email
                    fail_silently=False,
                )

            return Response(groupe_serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(groupe_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response(status=status.HTTP_405_METHOD_NOT_ALLOWED)


@api_view(['PUT'])
def update_groupe(request, groupe_id):
    try:
        groupe = FichePresnce.objects.get(id=groupe_id)
    except FichePresnce.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    groupe_data = request.data
    serializer = FichePresnceSerializer(groupe, data=groupe_data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def get_groupe_by_id(request, groupe_id):
    if request.method == 'GET':
        try:
            groupe = FichePresnce.objects.get(id=groupe_id)
        except FichePresnce.DoesNotExist:
            return Response({"message": "Groupe not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = FichePresnceSerializer(groupe)
        return Response(serializer.data, status=status.HTTP_200_OK)

    return Response({"message": "Method not allowed"}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

@api_view(['GET'])
def get_all_groupes(request):
    if request.method == 'GET':
        groupes = FichePresnce.objects.all()
        serializer = FichePresnceSerializer(groupes, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    else:
        return Response({"message": "Method not allowed"}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

@api_view(['GET'])
def get_classe_by_cin(request, cin):
    if request.method == 'GET':
        try:
            # Recherchez parmi tous les groupes pour trouver l'étudiant
            groupes = FichePresnce.objects.filter(Etudiants__Cin=cin)
            if groupes:
                # Supposons que l'étudiant est unique dans les groupes
                groupe = groupes[0]
                # Supposons que la classe est déduite à partir du niveau et de la spécialité
                classe = f"{groupe.Niveau} {groupe.Specialite}"
                return Response({"classe": classe}, status=status.HTTP_200_OK)
            else:
                return Response({"message": "Etudiant not found"}, status=status.HTTP_404_NOT_FOUND)
        except FichePresnce.DoesNotExist:
            return Response({"message": "Groupe not found"}, status=status.HTTP_404_NOT_FOUND)

    return Response({"message": "Method not allowed"}, status=status.HTTP_405_METHOD_NOT_ALLOWED)






@api_view(['GET'])
def get_groupes_by_cinenseignant(request, cinenseignant):
    if request.method == 'GET':
        try:
            # Recherchez parmi tous les groupes pour trouver ceux associés à cet enseignant
            groupes = FichePresnce.objects.filter(cinEnseignant=cinenseignant)
            if groupes:
                serializer = FichePresnceSerializer(groupes, many=True)
                # Construire une liste de dictionnaires contenant à la fois l'ID du groupe et les données du groupe
                groupes_data = [{'id': str(groupe.id), **serializer.data[i]} for i, groupe in enumerate(groupes)]
                return Response(groupes_data, status=status.HTTP_200_OK)
            else:
                return Response({"message": "Aucun groupe trouvé pour cet enseignant"}, status=status.HTTP_404_NOT_FOUND)
        except FichePresnce.DoesNotExist:
            return Response({"message": "Groupe non trouvé"}, status=status.HTTP_404_NOT_FOUND)

    return Response({"message": "Méthode non autorisée"}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

@api_view(['DELETE'])
def delete_groupe(request, groupe_id):
    try:
        groupe = FichePresnce.objects.get(id=groupe_id)
    except FichePresnce.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    groupe.delete()
    return Response({"message": "Groupe supprimé avec succès"}, status=status.HTTP_204_NO_CONTENT)




from django.http import JsonResponse
from django.shortcuts import get_object_or_404

import re

from datetime import datetime, time

@api_view(['GET'])
def allow_update(request, idFiche):
    if request.method == 'GET':
        try:
            # Récupérer la fiche de présence associée à l'ID fourni
            fiche = get_object_or_404(FichePresnce, id=idFiche)
            
            # Utiliser une expression régulière pour extraire les heures de début et de fin de la séance
            regex = r'(\d+)\s*:\s*(\d+)\s*-\s*(\d+)\s*:\s*(\d+)'

            match = re.match(regex, fiche.Seance)
            if match:
                # Extrait les heures de début et de fin de la séance
                start_hour, start_minute, end_hour, end_minute = map(int, match.groups())
                
                # Récupérer l'heure actuelle et la date
                current_time = datetime.now().time()
                current_date = datetime.now().date()
                
                # Créer une date à partir de l'heure de début de la séance
                seance_date = datetime.combine(fiche.DateSeance, time(start_hour, start_minute))
                
                # Comparer la date et l'heure actuelles avec la date et l'heure de fin de la séance
                if current_date == fiche.DateSeance and \
                   (current_time.hour < end_hour or \
                   (current_time.hour <= end_hour and current_time.minute < end_minute)):
                    return JsonResponse({
                        'current_datetime': str(datetime.combine(current_date, current_time)),
                        'seance_end_datetime': str(datetime.combine(current_date, time(end_hour, end_minute))),
                        'allowed': True
                    })
                else:
                    return JsonResponse({
                        'current_datetime': str(datetime.combine(current_date, current_time)),
                        'seance_end_datetime': str(datetime.combine(current_date, time(end_hour, end_minute))),
                        'allowed': False
                    })
            else:
                return JsonResponse({'error': 'Format de séance incorrect'}, status=400)
        except FichePresnce.DoesNotExist:
            return JsonResponse({'error': 'Fiche de présence non trouvée'}, status=404)
        except ValueError:
            return JsonResponse({'error': 'Format de séance incorrect'}, status=400)
    else:
        return JsonResponse({'error': 'Méthode non autorisée'}, status=405)



import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from .models import FichePresnce

@api_view(['GET'])
def export_fiche_presence_xlsx(request, fiche_id):
    try:
        # Récupérer la fiche de présence par ID
        fiche = FichePresnce.objects.get(id=fiche_id)
    except FichePresnce.DoesNotExist:
        return JsonResponse({'error': 'Fiche de présence non trouvée'}, status=404)

    # Créer un nouveau workbook et une feuille
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Fiche de Présence"

    # Ajouter des en-têtes
    headers = ["Nom de Groupe", "Séance", "Date de Séance", "Enseignant", "Salle", "Matière", "CIN Enseignant", "Année Universitaire"]
    student_headers = ["Nom Etudiant", "CIN", "Etat Absence", "Etat Elimination"]

    for col_num, header in enumerate(headers + student_headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Ajouter les données de la fiche
    fiche_data = [
        fiche.nomGroupe, fiche.Seance, fiche.DateSeance.strftime('%Y-%m-%d'), fiche.enseignant, 
        fiche.Sale, fiche.Matiere, fiche.cinEnseignant, fiche.Anneeuniversitaire
    ]
    
    for col_num, data in enumerate(fiche_data, 1):
        sheet.cell(row=2, column=col_num).value = data

    # Ajouter les données des étudiants
    for row_num, etudiant in enumerate(fiche.Etudiants, 3):
        etudiant_data = [
            etudiant.NomEtudiant, etudiant.Cin, etudiant.etatAbscence, etudiant.etatElimination
        ]
        for col_num, data in enumerate(etudiant_data, 1):
            sheet.cell(row=row_num, column=len(headers) + col_num).value = data

    # Préparer la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Fiche_Presence_{fiche.nomGroupe}_{fiche.DateSeance.strftime("%Y-%m-%d")}.xlsx'

    # Enregistrer le workbook dans la réponse
    workbook.save(response)
    return response



from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from .models import FichePresnce

@api_view(['GET'])
def export_fiche_presence_pdf(request, fiche_id):
    try:
        # Récupérer la fiche de présence par ID
        fiche = FichePresnce.objects.get(id=fiche_id)
    except FichePresnce.DoesNotExist:
        return JsonResponse({'error': 'Fiche de présence non trouvée'}, status=404)

    # Créer une réponse HTTP avec le type de contenu PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=Fiche_Presence_{fiche.nomGroupe}_{fiche.DateSeance.strftime("%Y-%m-%d")}.pdf'

    # Créer un objet canvas
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    # Définir une marge
    margin = inch

    # Ajouter un titre
    p.setFont("Helvetica-Bold", 16)
    p.drawString(margin, height - margin, "Fiche de Présence")

    # Ajouter les informations de la fiche de présence
    p.setFont("Helvetica", 12)
    fiche_info = [
        f"Nom de Groupe: {fiche.nomGroupe}",
        f"Séance: {fiche.Seance}",
        f"Date de Séance: {fiche.DateSeance.strftime('%Y-%m-%d')}",
        f"Enseignant: {fiche.enseignant}",
        f"Salle: {fiche.Sale}",
        f"Matière: {fiche.Matiere}",
        f"CIN Enseignant: {fiche.cinEnseignant}",
        f"Année Universitaire: {fiche.Anneeuniversitaire}"
    ]

    y_position = height - margin - 40
    for info in fiche_info:
        p.drawString(margin, y_position, info)
        y_position -= 20

    # Ajouter les en-têtes des étudiants
    p.setFont("Helvetica-Bold", 12)
    student_headers = ["Nom Etudiant", "CIN", "Etat Absence", "Etat Elimination"]
    x_position = margin
    for header in student_headers:
        p.drawString(x_position, y_position, header)
        x_position += 1.5 * inch

    # Ajouter les données des étudiants
    p.setFont("Helvetica", 12)
    y_position -= 20
    for etudiant in fiche.Etudiants:
        x_position = margin
        etudiant_data = [etudiant.NomEtudiant, etudiant.Cin, etudiant.etatAbscence, etudiant.etatElimination]
        for data in etudiant_data:
            p.drawString(x_position, y_position, data)
            x_position += 1.5 * inch
        y_position -= 20

    # Fermer le document PDF
    p.showPage()
    p.save()
    return response
