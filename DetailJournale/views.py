from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status

from Evaluation.models import EvaluationJournale
from Soutenance.models import Soutenance
from Tache.models import Tache
from .models import DetailJournale
from .Serializers import DetailJournaleSerializer
from django.http import JsonResponse

@api_view(['POST'])
def create_Journale(request):
    if request.method == 'POST':
        journale_serializer = DetailJournaleSerializer(data=request.data)
        if journale_serializer.is_valid():
            journale_serializer.save()
            return Response(journale_serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(journale_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    else:
        return Response(status=status.HTTP_405_METHOD_NOT_ALLOWED)

@api_view(['PUT'])
def update_Journale(request, groupe_id):
    try:
        journale = DetailJournale.objects.get(id=groupe_id)
    except DetailJournale.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    groupe_data = request.data
    serializer = DetailJournaleSerializer(journale, data=groupe_data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
def get_Journale_by_id(request, groupe_id):
    if request.method == 'GET':
        try:
            journale = DetailJournale.objects.get(id=groupe_id)
        except DetailJournale.DoesNotExist:
            return Response({"message": "Groupe not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = DetailJournaleSerializer(journale)
        return Response(serializer.data, status=status.HTTP_200_OK)

    return Response({"message": "Method not allowed"}, status=status.HTTP_405_METHOD_NOT_ALLOWED)


@api_view(['GET'])
def get_all_journale(request):
    if request.method == 'GET':
        journales = DetailJournale.objects.all()
        serialized_journales = []
        for journale in journales:
            journale_data = {
                'id': str(journale.id),  # Conversion en chaîne pour l'ID si nécessaire
                'anneeUniversitaire': journale.anneeUniversitaire,
                'Nom': journale.Nom,
                'Cin': journale.Cin,
                'Prenom': journale.Prenom,
                'Departement': journale.Departement,
                'Option': journale.Option,
                'Entreprise': journale.Entreprise,
                'Adress': journale.Adress,
                'Fax': journale.Fax,
                'NatureStage': journale.NatureStage,
                'PeriodeStage': journale.PeriodeStage,
                'etatEvaluation': journale.etatEvaluation,
                
            }
            serialized_journales.append(journale_data)
        return Response(serialized_journales, status=status.HTTP_200_OK)
    else:
        return Response({"message": "Method not allowed"}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

@api_view(['GET'])
def mon_journale(request, cin , id):
    if request.method == 'GET':
        journales = DetailJournale.objects.filter(Cin=cin , id=id)
        serialized_journales = []
        for journale in journales:
            journale_data = {
                'id': str(journale.id),  # Conversion en chaîne pour l'ID si nécessaire
                'anneeUniversitaire': journale.anneeUniversitaire,
                'Nom': journale.Nom,
                'Cin': journale.Cin,
                'Prenom': journale.Prenom,
                'Departement': journale.Departement,
                'Option': journale.Option,
                'Entreprise': journale.Entreprise,
                'Adress': journale.Adress,
                'Fax': journale.Fax,
                'NatureStage': journale.NatureStage,
                'PeriodeStage': journale.PeriodeStage,
                'etatEvaluation': journale.etatEvaluation,
                
            }
            
        return Response(journale_data, status=status.HTTP_200_OK)
    else:
        return Response({"message": "Method not allowed"}, status=status.HTTP_405_METHOD_NOT_ALLOWED)


@api_view(['DELETE'])
def delete_journale(request, pk):
    try:
        actualite = DetailJournale.objects.get(pk=pk)
    except DetailJournale.DoesNotExist:
        return JsonResponse({"error": "journale not found"}, status=404)
    
    if request.method == 'DELETE':
        actualite.delete()
        return JsonResponse({"message": "journale deleted successfully"}, status=200)


@api_view(['GET'])
def mes_journales(request, Cin):
    if request.method == 'GET':
        journales = DetailJournale.objects.filter(Cin=Cin)
        serialized_journales = []
        for journale in journales:
            journale_data = {
                'id': str(journale.id),  # Conversion en chaîne pour l'ID si nécessaire
                'anneeUniversitaire': journale.anneeUniversitaire,
                'Nom': journale.Nom,
                'Cin': journale.Cin,
                'Prenom': journale.Prenom,
                'Departement': journale.Departement,
                'Option': journale.Option,
                'Entreprise': journale.Entreprise,
                'Adress': journale.Adress,
                'Fax': journale.Fax,
                'NatureStage': journale.NatureStage,
                'PeriodeStage': journale.PeriodeStage,
                'etatEvaluation': journale.etatEvaluation,
            }
            serialized_journales.append(journale_data)
            
        return Response(serialized_journales, status=status.HTTP_200_OK)
    else:
        return Response({"message": "Method not allowed"}, status=status.HTTP_405_METHOD_NOT_ALLOWED)


@api_view(['PUT'])
def evaluer(request, groupe_id):
    try:
        journale = DetailJournale.objects.get(id=groupe_id)
    except DetailJournale.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    # Update the etatEvaluation field to True
    journale.etatEvaluation = 'Évalué'
    journale.save()

    serializer = DetailJournaleSerializer(journale)
    return Response(serializer.data, status=status.HTTP_200_OK)



@api_view(['POST'])
def get_journales_by_cins(request):
    if request.method == 'POST':
        cins = request.data.get('cins', [])
        
        if not isinstance(cins, list):
            return Response({"error": "Invalid data format. 'cins' should be a list."}, status=status.HTTP_400_BAD_REQUEST)
        
        journales = DetailJournale.objects.filter(Cin__in=cins)
        serialized_journales = DetailJournaleSerializer(journales, many=True).data
        
        return Response(serialized_journales, status=status.HTTP_200_OK)
    
    return Response({"message": "Method not allowed"}, status=status.HTTP_405_METHOD_NOT_ALLOWED)



@api_view(['GET'])
def get_students_by_president_jury_cin(request, cin_president_jury):
    try:
        # Filtrer les soutenances par le CIN du président du jury
        soutenances = Soutenance.objects.filter(CinPresidentJuri=cin_president_jury)

        # Vérifier si des soutenances ont été trouvées
        if soutenances:
            # Extraire les CIN des étudiants associés à ces soutenances
            students_cins = set()
            for soutenance in soutenances:
                for etudiant in soutenance.Etudiants:
                    students_cins.add(etudiant.Cin)

            # Récupérer les journaux des étudiants
            journales = DetailJournale.objects.filter(Cin__in=students_cins)
            serialized_journales = []
            for journale in journales:
              journale_data = {
                    'id': str(journale.id),  # Conversion en chaîne pour l'ID si nécessaire
                    'anneeUniversitaire': journale.anneeUniversitaire,
                    'Nom': journale.Nom,
                    'Cin': journale.Cin,
                    'Prenom': journale.Prenom,
                    'Departement': journale.Departement,
                    'Option': journale.Option,
                    'Entreprise': journale.Entreprise,
                    'Adress': journale.Adress,
                    'Fax': journale.Fax,
                    'NatureStage': journale.NatureStage,
                    'PeriodeStage': journale.PeriodeStage,
                    'etatEvaluation': journale.etatEvaluation,
                }
            serialized_journales.append(journale_data)
            
            return Response(serialized_journales, status=status.HTTP_200_OK)
        else:
            return Response({"message": "Aucune soutenance trouvée pour ce président de jury"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def get_students_by_encadreur_cin(request, cin_president_jury):
    try:
        # Filtrer les soutenances par le CIN du président du jury
        soutenances = Soutenance.objects.filter(CinEncadreur=cin_president_jury)

        # Vérifier si des soutenances ont été trouvées
        if soutenances:
            # Extraire les CIN des étudiants associés à ces soutenances
            students_cins = set()
            for soutenance in soutenances:
                for etudiant in soutenance.Etudiants:
                    students_cins.add(etudiant.Cin)

            # Récupérer les journaux des étudiants
            journales = DetailJournale.objects.filter(Cin__in=students_cins)
            serialized_journales = []
            for journale in journales:
              journale_data = {
                    'id': str(journale.id),  # Conversion en chaîne pour l'ID si nécessaire
                    'anneeUniversitaire': journale.anneeUniversitaire,
                    'Nom': journale.Nom,
                    'Cin': journale.Cin,
                    'Prenom': journale.Prenom,
                    'Departement': journale.Departement,
                    'Option': journale.Option,
                    'Entreprise': journale.Entreprise,
                    'Adress': journale.Adress,
                    'Fax': journale.Fax,
                    'NatureStage': journale.NatureStage,
                    'PeriodeStage': journale.PeriodeStage,
                    'etatEvaluation': journale.etatEvaluation,
                }
            serialized_journales.append(journale_data)
            
            return Response(serialized_journales, status=status.HTTP_200_OK)
        else:
            return Response({"message": "Aucune soutenance trouvée pour ce président de jury"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



@api_view(['GET'])
def get_students_by_rapporteur_cin(request, cin_president_jury):
    try:
        # Filtrer les soutenances par le CIN du président du jury, du rapporteur ou de l'encadreur
        soutenances = Soutenance.objects.filter(
            CinRaporteur=cin_president_jury
        ) | Soutenance.objects.filter(
            CinEncadreur=cin_president_jury
        ) | Soutenance.objects.filter(
            CinPresidentJury=cin_president_jury
        )

        # Vérifier si des soutenances ont été trouvées
        if soutenances.exists():
            # Extraire les CIN des étudiants associés à ces soutenances
            students_cins = set()
            for soutenance in soutenances:
                for etudiant in soutenance.Etudiants.all():
                    students_cins.add(etudiant.Cin)

            # Récupérer les journaux des étudiants
            journales = DetailJournale.objects.filter(Cin__in=students_cins)
            serialized_journales = DetailJournaleSerializer(journales, many=True).data

            return Response(serialized_journales, status=status.HTTP_200_OK)
        else:
            return Response({"message": "Aucune soutenance trouvée pour ce président de jury"}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    


from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .models import  DetailJournale
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from openpyxl.styles import Font
from django.http import HttpResponse
from rest_framework.decorators import api_view
from .models import DetailJournale

@api_view(['GET'])
def export_all_data_xlsx(request, detail_id):
    try:
        detail = DetailJournale.objects.get(id=detail_id)
    except DetailJournale.DoesNotExist:
        return JsonResponse({'error': 'Détail journalier non trouvé'}, status=404)

    # Créer un nouveau workbook et une feuille
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Détails Journalier"

    # Ajouter des en-têtes pour DetailJournale
    headers = [
        "Année Universitaire", "Nom", "CIN", "Prénom", "Département", "Option",
        "Entreprise", "Adresse", "Fax", "Nature du Stage", "Période du Stage", "État de l'Évaluation"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Ajouter les données de DetailJournale
    detail_data = [
        detail.anneeUniversitaire, detail.Nom, detail.Cin, detail.Prenom,
        detail.Departement, detail.Option, detail.Entreprise, detail.Adress,
        detail.Fax, detail.NatureStage, detail.PeriodeStage, detail.etatEvaluation
    ]
    for col_num, data in enumerate(detail_data, 1):
        sheet.cell(row=2, column=col_num).value = data

    # Créer une nouvelle feuille pour les Taches
    tache_sheet = workbook.create_sheet(title="Tâches")
    tache_headers = ["Date", "Tâche Journalière", "Dernière Modification"]
    for col_num, header in enumerate(tache_headers, 1):
        cell = tache_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Ajouter les données des Taches
    taches = Tache.objects.filter(idJournale=str(detail.id))
    for row_num, tache in enumerate(taches, 2):
        tache_data = [tache.Date.strftime('%Y-%m-%d'), tache.TacheJournaliere, tache.LastModified.strftime('%Y-%m-%d')]
        for col_num, data in enumerate(tache_data, 1):
            tache_sheet.cell(row=row_num, column=col_num).value = data

    # Créer une nouvelle feuille pour les Evaluations
    eval_sheet = workbook.create_sheet(title="Évaluations")
    eval_headers = [
        "Nom du Rapporteur", "Forme d'Expression", "Forme de Présentation", "Présentation de l'Entreprise",
        "Valeur Scientifique", "Effort Personnel", "Documentation", "Contact avec l'Entreprise", "Observation"
    ]
    for col_num, header in enumerate(eval_headers, 1):
        cell = eval_sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)

    # Ajouter les données des Evaluations
    evaluations = EvaluationJournale.objects.filter(idJournale=str(detail.id))
    for row_num, evaluation in enumerate(evaluations, 2):
        eval_data = [
            evaluation.NomRapporteur, evaluation.FormeExpression, evaluation.FormePesentation, evaluation.PresentationEntreprise,
            evaluation.ValeurScien, evaluation.EffortPersonnel, evaluation.Documentation, evaluation.ContactEntreprise, evaluation.Observation
        ]
        for col_num, data in enumerate(eval_data, 1):
            eval_sheet.cell(row=row_num, column=col_num).value = data

    # Préparer la réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Détails_{detail.Nom}_{detail.Cin}.xlsx'
    workbook.save(response)
    return response



from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas

@api_view(['GET'])
def export_all_data_pdf(request, detail_id):
    try:
        detail = DetailJournale.objects.get(id=detail_id)
    except DetailJournale.DoesNotExist:
        return JsonResponse({'error': 'Détail journalier non trouvé'}, status=404)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=Détails_{detail.Nom}_{detail.Cin}.pdf'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    margin = inch

    p.setFont("Helvetica-Bold", 16)
    p.drawString(margin, height - margin, "Détail Journalier")

    p.setFont("Helvetica", 12)
    detail_info = [
        f"Année Universitaire: {detail.anneeUniversitaire}",
        f"Nom: {detail.Nom}",
        f"CIN: {detail.Cin}",
        f"Prénom: {detail.Prenom}",
        f"Département: {detail.Departement}",
        f"Option: {detail.Option}",
        f"Entreprise: {detail.Entreprise}",
        f"Adresse: {detail.Adress}",
        f"Fax: {detail.Fax}",
        f"Nature du Stage: {detail.NatureStage}",
        f"Période du Stage: {detail.PeriodeStage}",
        f"État de l'Évaluation: {detail.etatEvaluation}"
    ]

    y_position = height - margin - 40
    for info in detail_info:
        p.drawString(margin, y_position, info)
        y_position -= 20

    p.setFont("Helvetica-Bold", 14)
    p.drawString(margin, y_position - 20, "Tâches")
    y_position -= 40
    p.setFont("Helvetica", 12)
    taches = Tache.objects.filter(idJournale=str(detail.id))
    for tache in taches:
        tache_info = f"Date: {tache.Date.strftime('%Y-%m-%d')} - Tâche Journalière: {tache.TacheJournaliere} - Dernière Modification: {tache.LastModified.strftime('%Y-%m-%d')}"
        p.drawString(margin, y_position, tache_info)
        y_position -= 20

    p.setFont("Helvetica-Bold", 14)
    p.drawString(margin, y_position - 20, "Évaluations")
    y_position -= 40
    p.setFont("Helvetica", 12)
    evaluations = EvaluationJournale.objects.filter(idJournale=str(detail.id))
    for evaluation in evaluations:
        eval_info = [
            f"Nom du Rapporteur: {evaluation.NomRapporteur}",
            f"Forme d'Expression: {evaluation.FormeExpression}",
            f"Forme de Présentation: {evaluation.FormePesentation}",
            f"Présentation de l'Entreprise: {evaluation.PresentationEntreprise}",
            f"Valeur Scientifique: {evaluation.ValeurScien}",
            f"Effort Personnel: {evaluation.EffortPersonnel}",
            f"Documentation: {evaluation.Documentation}",
            f"Contact avec l'Entreprise: {evaluation.ContactEntreprise}",
            f"Observation: {evaluation.Observation}"
        ]
        for info in eval_info:
            p.drawString(margin, y_position, info)
            y_position -= 20
        y_position -= 20  # Ajouter un espace entre chaque évaluation

    p.showPage()
    p.save()
    return response
